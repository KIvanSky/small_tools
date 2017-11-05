import openpyxl

import os

os.mkdir('/Users/BuleSky/Desktop/test') # 结果文件夹路径

file_dir = "/Users/BuleSky/Desktop/data"    # 数据文件夹路径
for root, dirs, files in os.walk(file_dir):
    for i in range(4):  # 文件夹里几个文件范围就是几

        name = files[i] # windows应该不需要+1
        print(name)
        inwb = openpyxl.load_workbook('/Users/BuleSky/Desktop/data/'+name)
        print('end')
        sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
        ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
        rows = ws.max_row  # 获取读取的excel的文件的行数
        column = ws.max_column
        ws.cell(row=1, column=column + 1).value = '液位波动'
        # print(ws.cell(row=1, column=column + 1).value)
        for k in range(rows-1):
            num1 = ws.cell(row=k+2, column=25).value
            num2 = ws.cell(row=k+2, column=26).value
            # print(num1, num2)

            result = 160 * (float(num1) / 10000) - 160 * (float(num2) / 10000)
            ws.cell(row=k+2, column=column + 1).value = result
            print(ws.cell(row=k+2, column=column + 1).value)
        os.chdir('/Users/BuleSky/Desktop/test')
        outwb = openpyxl.Workbook()
        outws = outwb.get_active_sheet()
        for r in range(rows):
            for c in range(column+1):
                outws.cell(row=r+1, column=c+1).value = ws.cell(row=r+1, column=c+1).value  # 读文件
                # print(outws.cell(row=r+1, column=c+1).value)
        new_name = name[0:-5]
        outwb.save(new_name+'1.xlsx')  # 一定要记得保存
        print(rows, column)





